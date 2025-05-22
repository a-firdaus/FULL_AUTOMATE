# This python script aims to extract the bader charge from file based on the index of
# atom selected in a POSCAR file and write it in an excel sheet that already exists.
# Note 1: Be sure to adapt the 'excel_positions' variable to write the charges at the right
#         place in your own excel sheet
# Note 2 : the excel sheet cannot be modified while it is opened
# Note 3 : the script assumes the charge written in the file is atomic charge and not bader, as a consequence,
#          it subtracts this charge to ZVAL taken from the dictionary

from math import sqrt
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
from vpython import vector as vpythvector
import tkinter as tk
from tkinter import filedialog, simpledialog
from matplotlib.widgets import LassoSelector
from matplotlib.path import Path
import pyautogui
from matplotlib.lines import Line2D
import openpyxl
from molatomclasses import atom
import re
matplotlib.use('TkAgg')

radiuslist={"H":1.20,"Li":1.82,"Cl":2.75 ,"C":1.70, "Si":2.10 ,"N":1.55, "O":1.52,"P":1.80,"S":1.80,
            'Ne':1.54,'Ar':1.88,'He':1.40,"Na":2.27,"F":1.47,"Br":1.85,"Tc":2.05,"Pm":2.43,"Cd":1.55,
            "Co":1.35,"Mn":1.61,"Fe":1.61,"Ni":1.49}
colourdict = {"H": (0.8, 0.8, 0.8), "Li":(252/255, 144/255, 245/255),"Cl": (0, 1, 0), "C": (0, 0, 0), "O": (1, 0, 0), "N": (0, 0, 1), "Ar": (0.90, 0.90, 0.90),
              "Ne": (0.90, 0.90, 0.90), "S": (1, 1, 0), "He": (0.90, 0.90, 0.90), "Na": (1, 1, 1), "Pm": (1, 0.8, 1),
              "Tc": (0.8, 1, 1), "F": (159 / 255.0, 252 / 255.0, 206 / 255.0), "Br": (153 / 255.0, 34 / 255.0, 0),
              "Cd": (0.8, 1, 1), "Co": (0.8, 1, 1),"Si":(153/255, 51/255, 242/255),"Ni":(0.5, 1, 1),"Mn":(0.3, 1, 1),"P":(180/255.0, 52/255.0, 235/255.0),"Fe":(0.3, 1, 1)}
molarweightlist={"H":1.00794,"O":15.9994,"C":12.0107,"Ar":39.948,"He":4.002602,"Ne":20.1797,"N":14.0067,
                 "Cl":35.453,"Na":22.989769,"F":18.998403,"Br":79.904,"Tc":98.0,"Pm":145.0,"Cd":112.411,
                 "Co":58.9332,"fulvic_acid":308.242,"Pb":207.2,'Cu':63.546,"Fe":55.845}
element_colors = {'H': 'oldlace', 'Li': 'green', 'O': 'red', 'C': 'saddlebrown', 'N': 'blue',
                      'Mn': 'darkorchid', 'Ni': 'silver', 'Co': 'indigo'}
excel_positions = {
    'LiNMC811':{'Alcohol':(2,5),
                               'Amine':(5,5),
                               'Carbonate': {"Simple": (8,5), "Double": (12, 5)},
                               'Ester':{"Simple": (14,5), "Double": (17, 5)},
                               'Ether':(20,5),
                               'Urethane':{"Simple": (23,5), "Double": (25, 5), "Nitrogen": (27, 5)}},
    'LiNMC622': {'Alcohol':(29,5),
                               'Amine':(32,5),
                               'Carbonate': {"Simple": (35,5), "Double": (39, 5)},
                               'Ester':{"Simple": (41,5), "Double": (44, 5)},
                               'Ether':(47,5),
                               'Urethane':{"Simple": (51,5), "Double": (53, 5), "Nitrogen": (55, 5)}},
    'NMC811':  {'Alcohol':(56,5),
                               'Amine':(59,5),
                               'Carbonate': {"Simple": (62,5), "Double": (66, 5)},
                               'Ester':{"Simple": (68,5), "Double": (71, 5)},
                               'Ether':(74,5),
                               'Urethane':{"Simple": (77, 5), "Double": (79, 5), "Nitrogen": (81, 5)}},
    'NMC622':  {'Alcohol':(83,5),
                               'Amine':(86,5),
                               'Carbonate': {"Simple": (89,5), "Double": (93, 5)},
                               'Ester':{"Simple": (95,5), "Double": (98, 5)},
                               'Ether':(101,5),
                               'Urethane':{"Simple": (104,5), "Double": (106, 5), "Nitrogen": (108, 5)}}} # Adapt to your system
valences = {'C': 4, 'Co': 9, 'H': 1, 'Li': 1, 'Mn': 7, 'N': 5, 'Ni': 10, 'O': 6} # Adapt to your system


def euclidiandistance_vectors(pos1,pos2):
    return sqrt(((pos2.x-pos1.x)**2)+((pos2.y-pos1.y)**2)+((pos2.z-pos1.z)**2))
class SelectFromCollection(object):
    def __init__(self, ax, collection, alpha_other=0.15):
        self.canvas = ax.figure.canvas
        self.collection = collection
        self.alpha_other = alpha_other
        self.xys = collection.get_offsets()
        self.Npts = len(self.xys)
        self.fc = collection.get_facecolors()
        if len(self.fc) == 0:
            raise ValueError('Collection must have a facecolor')
        elif len(self.fc) == 1:
            self.fc = np.tile(self.fc, (self.Npts, 1))
        self.lasso = LassoSelector(ax, onselect=self.onselect)
        self.ind = []
        self.canvas.get_tk_widget().focus_set()
    def onselect(self, verts):
        path = Path(verts)
        self.ind = np.nonzero(path.contains_points(self.xys))[0]
        self.fc[:, -1] = self.alpha_other
        self.fc[self.ind, -1] = 1
        self.collection.set_facecolors(self.fc)
        self.canvas.get_tk_widget().focus_set()
        self.canvas.draw_idle()
    def disconnect(self):
        self.lasso.disconnect_events()
        self.fc[:, -1] = 1
        self.collection.set_facecolors(self.fc)
        self.canvas.draw_idle()
def accept(event, selector, all_selections):
    if event.key == "enter":
        print("Point selection OK!\nHere are the indices:")
        #all_selections.append(selector.ind.tolist()[0])
        all_selections.append(selector.ind)
        print(all_selections)
        selector.disconnect()
        ax.set_title("(Selection OK, please close the window)")
        fig.canvas.draw()

def draw(molecule, atoms, radiuslist, element_colors):
    orientation_option = pyautogui.confirm(
        text="Select the appropriate axis view.\nIf this is not the right choice, then press cancel in a later dialog to return here.",
        title="Select the axis orientation", buttons=["XY", "YZ", "XZ"])
    data_firstaxis = []
    data_secondaxis = []
    sizes = []
    colours = []
    for atomi in molecule:
        if orientation_option == "XY":
            data_firstaxis.append(atomi.pos.x)
            data_secondaxis.append(atomi.pos.y)
            sizes.append(radiuslist[atomi.symbol] * 20)
            colours.append(element_colors[atomi.symbol])
        elif orientation_option == "YZ":
            data_firstaxis.append(atomi.pos.z)
            data_secondaxis.append(atomi.pos.y)
            sizes.append(radiuslist[atomi.symbol] * 20)
            colours.append(element_colors[atomi.symbol])
        elif orientation_option == "XZ":
            data_firstaxis.append(atomi.pos.x)
            data_secondaxis.append(atomi.pos.z)
            sizes.append(radiuslist[atomi.symbol] * 20)
            colours.append(element_colors[atomi.symbol])
        else:
            raise NotImplementedError
    ax = plt.gca()
    fig = plt.gcf()
    pts = ax.scatter(data_firstaxis, data_secondaxis, color=colours, s=sizes)
    # Set the background color to a fading grey gradient
    # ax.set_facecolor((0.9, 0.9, 0.9, 0.5))  # Fading grey with alpha blending
    # Lists to store legend handles and labels
    marker_handles = []
    marker_labels = []
    # Add legend handles and labels for markers
    for element in atoms:
        if element not in marker_labels:
            marker_handles.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=element_colors[element],
                                         markersize=8, label=element))  # Adjust the markersize value as needed
            marker_labels.append(element)
        # Create custom legends for materials and markers
        marker_legend = plt.legend(handles=marker_handles, labels=marker_labels, title='Element', loc='upper left',
                                   bbox_to_anchor=(1.0, 1), fontsize=10, markerscale=1.5, title_fontsize=12)
    ax.set_title("Select the molecule and press enter to accept selected points. Otherwise, close the window.")
    return ax, fig, pts
def read_atom_positions(POSCARpath):
    root = tk.Tk()
    root.withdraw()
    atoms = []
    quantities = []
    history = []
    nbr_line = -9
    check = False
    avector = None
    bvector = None
    cvector = None
    latticevectorarray = None
    root.destroy()
    currentindex = -1
    allatompos = {}
    molecule = []

    with open(POSCARpath, 'r') as POSCARfile:
        linecount = 0
        for line in POSCARfile:
            nbr_line += 1
            linecount += 1
            if line.strip() == "Selective dynamics":
                check = True
                atoms = history[-2].split()
                quantities = [int(x) for x in history[-1].split()]
                atomsymbolfulllist = create_atomsymbolfulllist(atoms, quantities)

            if not check:
                history.append(line.strip())

            if linecount == 3:
                avector = np.array([float(k) for k in line.strip().split()]) / 10
                print("a= ", avector, "")
                continue
            elif linecount == 4:
                bvector = np.array([float(k) for k in line.strip().split()]) / 10
                print("b= ", bvector, "")
                continue
            elif linecount == 5:
                cvector = np.array([float(k) for k in line.strip().split()]) / 10
                print("c= ", cvector, "")
                latticevectorarray = np.stack([avector, bvector, cvector], axis=0)
                print("latticevectorarray=", latticevectorarray, "")
                continue
            elif linecount > 9 and linecount < sum(quantities, 10):
                currentindex += 1
                pos = line.strip().split()
                location_ = vpythvector(*[float(pos[0]), float(pos[1]), float(pos[2])])
                currentsymbol = atomsymbolfulllist[currentindex] + str(currentindex)
                allatompos[currentsymbol] = location_
        add_atomi_to_molecule(molecule, allatompos)

    print("Number of lines in the POSCAR file : ", nbr_line)
    print("Atoms: " + str(atoms))
    print("Quantities: " + str(quantities))
    a = np.linalg.norm(avector) * 10
    b = np.linalg.norm(bvector) * 10
    c = np.linalg.norm(cvector) * 10
    volume = a*b*c
    print("a length", a, "")
    print("b length", b, "")
    print("c length", c, "")
    print(atomsymbolfulllist)

    return molecule, atoms, atomsymbolfulllist, quantities, volume
def add_atomi_to_molecule(molecule, allatompos):
    for atomid, atomposlocation in allatompos.items():
        molecule.append(atom(atomid, ''.join([i for i in atomid if not i.isdigit()]), atomposlocation))
def create_atomsymbolfulllist(atoms, quantities):
    atomsymbolfulllist = []
    for index_ in range(len(atoms)):
        for j in range(quantities[index_]):
            atomsymbolfulllist.append(atoms[index_])
    return atomsymbolfulllist

"""
    First plot system from POSCAR file
    Create a selection of indices --> as many selection as different types of atom of interst
"""

if __name__ == "__main__":
    POSCARpath = filedialog.askopenfilename(title="Select the POSCAR file")
    molecule, atoms, atomsymbolfulllist, quantities, volume = read_atom_positions(POSCARpath)
    has_been_selected = False
    all_selections = []
    selectionMean = str(simpledialog.askstring("Selection", "Do you want to use a lasso selector "
                                                            "to select atoms?"))
    lassoSelector = True
    if selectionMean.upper() == "NO":
        lassoSelector = False
    if lassoSelector:
        while not has_been_selected:
            ax, fig, pts = draw(molecule, atoms, radiuslist, element_colors)
            fig.canvas.get_tk_widget().focus_set()
            selector = SelectFromCollection(ax, pts)
            fig.canvas.mpl_connect("key_press_event", lambda event: accept(event, selector, all_selections))
            ax.set_title("Select the ATOM and press enter to accept selected point. Otherwise, close the window.")
            plt.show()
            plt.close()

            if pyautogui.confirm(text="Do you still need to select atoms?", buttons=["Yes", "No"]) == "Yes":
                continue
            else:
                has_been_selected = True

                bader_file = filedialog.askopenfilename(title="Select the file containing bader charges")
                excel_file = os.path.join(os.path.expanduser('~'), 'Desktop', 'SPE_NMC', 'EFIELD_simulations', 'Bader_Distance.xlsx') # Adapt to your system
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook['With EF'] # Adapt
                electrode = str(simpledialog.askstring("Material", "What is the electrode material?"))
                fg = str(simpledialog.askstring("Functional Group", "What is the functional group?"))
                type = None


                bader_charge = 0
                excel_cells = []
                with open(bader_file, 'r') as text_file:
                    count = 0
                    for i, ind in enumerate(all_selections):

                        if fg.upper() == 'CARBONATE':
                            type = str(simpledialog.askstring("Atom of interst",
                                                              "There are more than one atom of interst in this functional group. Choose one among: Simple or Double"))
                            excl_line, excl_col = excel_positions[electrode][fg][type]

                            cell_value = worksheet.cell(row=excl_line, column=excl_col).value
                            cell_is_empty = False

                        elif fg.upper() == "ESTER":
                            type = str(simpledialog.askstring("Atom of interst",
                                                              "There are more than one atom of interst in this functional group. Choose one among: Simple or Double"))
                            excl_line, excl_col = excel_positions[electrode][fg][type]

                            cell_value = worksheet.cell(row=excl_line, column=excl_col).value
                            cell_is_empty = False

                        elif fg.upper() == 'URETHANE':
                            type = str(simpledialog.askstring("Atom of interest",
                                                              "There are more than one atom of interest in this functional group. Choose one among: Simple, Double or Nitrogen"))
                            excl_line, excl_col = excel_positions[electrode][fg][type]

                            cell_value = worksheet.cell(row=excl_line, column=excl_col).value
                            cell_is_empty = False

                        else:
                            excl_line, excl_col = excel_positions[electrode][fg]

                            cell_value = worksheet.cell(row=excl_line, column=excl_col).value
                            cell_is_empty = False

                        if i == 0:
                            previous = type

                        print('-' * 50)
                        print("i:", i)
                        print("Index:",ind)
                        text_file.seek(0)
                        for line_number, line in enumerate(text_file, start=-1):
                            if line_number == ind + 1:

                                print('line number ', line_number)

                                if type != previous:
                                    count = 0
                                    previous = type

                                elem = re.sub(r'[^a-zA-Z]', '', worksheet.cell(row=excl_line+count, column=3).value)
                                zval = valences[elem]
                                bader_charge = zval - float(line.split()[4])

                                print("To write in the sheet:")
                                print(electrode, fg, type, line_number, bader_charge)

                                print("Where:", excl_line+count, excl_col)

                                worksheet.cell(row=excl_line+count, column=excl_col).value = bader_charge
                                worksheet.cell(row=excl_line+count, column=excl_col-1).value = line_number
                                print('Excel row:')
                                print(worksheet.cell(row=excl_line + count, column=1).value,
                                      worksheet.cell(row=excl_line + count, column=2).value,
                                      worksheet.cell(row=excl_line + count, column=3).value,
                                      worksheet.cell(row=excl_line + count, column=4).value,
                                      worksheet.cell(row=excl_line + count, column=5).value)

                                count += 1

                workbook.save(excel_file)
                # print('Average bader charge for the selection: ', bader_charge)
                print("Finish")
                break

    else:

        for i in range(12):
            idx = int(simpledialog.askstring("Index selection", "Give the index of an atom to analyze"))
            all_selections.append(idx)

        bader_file = filedialog.askopenfilename(title="Select the file containing bader charges")
        excel_file = r"D:\OneDrive\Bureau\Paper\Bader Analysis.xlsx"
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook['TM Analysis']
        electrode = str(simpledialog.askstring("Material", "What is the electrode material?"))
        fg = str(simpledialog.askstring("Functional Group", "What is the functional group?"))
        type = None

        bader_charge = 0
        excel_cells = []
        with open(bader_file, 'r') as text_file:

            excl_line, excl_col = excel_positions[electrode][fg]

            for i, ind in enumerate(sorted(all_selections)):
                print(ind)
                text_file.seek(0)
                for line_number, line in enumerate(text_file, start=-1):
                    if line_number == ind + 1:
                        print('line number ', line_number)
                        print(line.split())
                        atom = str(line.split()[1])
                        index = ind
                        bader_charge = float(line.split()[6])
                        worksheet.cell(row=excl_line + i, column=excl_col - 2).value = atom
                        worksheet.cell(row=excl_line + i, column=excl_col - 1).value = index
                        worksheet.cell(row=excl_line + i, column=excl_col).value = bader_charge

        workbook.save(excel_file)
        # print('Average bader charge for the selection: ', bader_charge)
        print("Finish")