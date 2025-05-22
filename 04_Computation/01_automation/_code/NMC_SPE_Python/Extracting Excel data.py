import openpyxl
import pandas as pd
import os

def extract_excel_data(file):
    # Get user input for sheet name, starting row, and starting column
    sheet_name = input("Enter the sheet name: ")

    # Check if the sheet name exists in the Excel file
    try:
        xl = pd.ExcelFile(file)
        if sheet_name not in xl.sheet_names:
            raise ValueError(f"The sheet '{sheet_name}' does not exist in the Excel file.")
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file}' does not exist.")

    # Validate and get the starting row input
    while True:
        try:
            start_row = int(input("Enter the starting row: "))
            if start_row <= 0:
                raise ValueError("Starting row must be a positive integer.")
            break
        except ValueError:
            print("Invalid input. Please enter a valid integer for starting row.")

    # Validate and get the starting column input
    while True:
        try:
            start_col = int(input("Enter the starting column: "))
            if start_col <= 0:
                raise ValueError("Starting column must be a positive integer.")
            break
        except ValueError:
            print("Invalid input. Please enter a valid integer for starting column.")

    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]

    if sheet.max_row < start_row or sheet.max_column < start_col:
        print("The specified starting row or column does not exist in the sheet.")
        return
    data = []

    row_counter = 0
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
        row_counter += 1
        if '' in row: # check if any cell in the row is empty
            break
        string_part = ' '.join(row[:2]) # joining the two first column as a string
        split_string = string_part.split()
        name = split_string[0]
        bond = split_string[1]
        print('Row counter : ', row_counter)
        string_part = f"{bond} ({name})" # formatting the string as per the requirement
        row = [string_part] + [float(x) for x in row[2:5]] # combining the first two column as a string and converting the rest to float
        data.append(row)


    return data

def remove_digits_from_strings(data):
    new_data = []

    for row in data:
        new_row = []
        for element in row:
            if isinstance(element, str):
                new_element = ''.join(ch for ch in element if not ch.isdigit())
                new_row.append(new_element)
            else:
                new_row.append(element)
        new_data.append(new_row)

    return new_data

def bond_name_order(data):
    new_data = []

    for row in data:
        new_row = []
        for element in row:
            if isinstance(element, str):
                if 'H-O' in element:
                    new_element = element.replace('H-O', 'O-H')
                    new_row.append(new_element)

                elif 'H-N' in element:
                    new_element = element.replace('H-N', 'N-H')
                    new_row.append(new_element)

                else:
                    new_row.append(element)


            else:
                new_row.append(element)
        new_data.append(new_row)

    return new_data

def remove_bond_type(data):
    new_data = []

    for row in data:
        new_row = []
        for element in row:
            if isinstance(element, str):
                if 'O-H' in element:
                    new_element = 'O-H'
                    new_row.append(new_element)

                elif 'C-C' in element:
                    new_element = 'C-C'
                    new_row.append(new_element)

                elif 'C-H' in element:
                    new_element = 'C-H'
                    new_row.append(new_element)

                else:
                    new_row.append(element)


            else:
                new_row.append(element)
        new_data.append(new_row)

    return new_data


# Replace these values with your actual file path and tab name
file = os.path.join(os.path.expanduser('~'), 'Desktop', 'SPE_NMC', 'EFIELD_simulations', 'Results_EF_Corrected.xlsx')

excel_data = extract_excel_data(file)
data_without_digits = remove_digits_from_strings(excel_data)
data_without_digits_bondname_ordered = bond_name_order(data_without_digits)
final_version_data = remove_bond_type(data_without_digits_bondname_ordered)
for line in final_version_data:
    print(line, ',')
